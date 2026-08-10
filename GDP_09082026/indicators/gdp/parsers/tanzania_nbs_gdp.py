"""Parser for the NBS Tanzania quarterly GDP workbook (xlsx, base 2015).

The NBS GDP release ('Excel Q<n>_<year>_Eng.xlsx') carries a single sheet
'TABLES 1-4' with four vertically-stacked, TRANSPOSED tables (economic activities
across the COLUMNS, periods down the rows: the annual figures first, then the
quarterly ones):

  GDP by Economic Activity at 2015 prices, TZS Million   -> level, constant (2015)
  GDP by Economic Activity at 2015 Prices, % Growth      -> growth_yoy
  GDP by Economic Activity at Current Prices, TZS Million-> level, current
  GDP by Economic Activity Current Prices, % Share       -> share

Each block has a caption row, a header row (Year | Quarter | <19 activities> | All
industry at basic prices | Taxes on products | GDP at market prices), then the
data. The Year appears only on the first row of each year, so it is forward-
filled; a blank Quarter marks an annual row, 1-4 a quarter. A font quirk renders
capital P as Q ('GDQ', 'Qublic', 'Qrices'), fixed on the labels. Units TZS million
(levels) / percent. 'All industry at basic prices', 'Taxes on products' and 'GDP
at market prices' are the aggregates. Verified: Agriculture 2024 constant =
37,677,354.6 TZS million. Nothing derived.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def _spec(caption: str):
    c = caption.lower()
    if "growth" in c:
        return ("growth_yoy", "constant", "2015", "percent")
    if "share" in c:
        return ("share", "current", "", "percent")
    if "2015 price" in c:
        return ("level", "constant", "2015", "TZS million")
    if "current price" in c:
        return ("level", "current", "", "TZS million")
    return None


def _fix(label):
    # this workbook's font maps capital P -> Q; restore it on the labels
    return re.sub(r"\s+", " ", label.replace("Q", "P")).strip() if isinstance(label, str) else ""


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def parse(pdf_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(pdf_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    n = len(grid)
    caps = [i for i, row in enumerate(grid)
            if row and isinstance(row[0], str)
            and "gross domestic product" in row[0].lower()]
    rows = []
    for k, ci in enumerate(caps):
        spec = _spec(grid[ci][0])
        if spec is None:
            continue
        measure, basis, base, unit = spec
        end = caps[k + 1] if k + 1 < len(caps) else n
        # header row: has 'Year' in col0 and 'Quarter' in col1
        hdr = None
        for r in range(ci, min(ci + 4, end)):
            if (isinstance(grid[r][0], str) and grid[r][0].strip().lower() == "year"):
                hdr = r
                break
        if hdr is None:
            continue
        activities = {c: _fix(grid[hdr][c]) for c in range(2, len(grid[hdr]))
                      if isinstance(grid[hdr][c], str) and grid[hdr][c].strip()}
        year = None
        for r in range(hdr + 1, end):
            row = grid[r]
            if not row:
                continue
            yv = row[0] if len(row) > 0 else None
            if isinstance(yv, (int, float)) and 1990 <= yv <= 2100:
                year = int(yv)
            q = row[1] if len(row) > 1 else None
            if year is None:
                continue
            if isinstance(q, (int, float)) and 1 <= q <= 4:
                period, freq = f"{year}-Q{int(q)}", "quarterly"
            elif q in (None, "") and any(_num(row[c]) is not None for c in activities):
                period, freq = str(year), "annual"
            else:
                continue
            for c, label in activities.items():
                v = _num(row[c] if c < len(row) else None)
                if v is None:
                    continue
                low = label.lower()
                approach = "aggregate" if ("gdp at market" in low or "all industry" in low
                                           or "taxes on product" in low) else "production"
                rows.append({
                    "approach": approach, "category": label, "category_group": "",
                    "series_code": "", "geography": "National", "period": period,
                    "frequency": freq, "price_basis": basis,
                    "seasonal_adjustment": "nsa", "measure": measure,
                    "value": v, "unit": unit, "base_period": base,
                })
    if not rows:
        raise ValueError("no GDP rows parsed from NBS Tanzania workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "period", "measure", "price_basis"])
