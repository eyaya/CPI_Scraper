"""Parser for the LISGIS (Liberia) annual GDP workbooks (xlsx, base year 2016).

Two published workbooks are read together (production = primary, expenditure =
extra). Each has the same table set as sheets:
  GDP at CP (current levels) / GDP at KP (constant levels) / Structure of GDP
  (share %) / Real Growth Rates (%) / Contribution to growth (pp) / Deflators
  (index, 2016=100).
Layout: header row carries the years across columns (constant/growth years are
suffixed '*'/'**' for provisional/estimate — stripped); column 1 = ISIC/line
code, column 2 = label; years from column 3.

IMPORTANT: the LEVEL sheets (CP, KP) contain the whole activity breakdown TWICE —
first in million Liberian dollars (the 'million $L' block), then again converted
to US$ millions after a blank gap and a repeated header. The USD block is a
derived FX conversion, so we read only the first (LRD) block and stop at the
repeated header. The rate/index sheets are a single block. Growth, structure and
contribution are published already as percentages (no rescaling). Verified: GDP
at market prices 2016 = 304,799.9 m$L (GVA 285,173.2 + net taxes 19,626.6), real
growth and 2016=100 deflator consistent. Nothing derived or converted.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

_YEAR_RE = re.compile(r"^(20\d\d)[\*\s]*$")


def _sheet_spec(name: str):
    """(measure, price_basis, base_period, unit) or None to skip the sheet."""
    n = name.lower()
    if "gdp at cp" in n or "gdp at current" in n:
        return ("level", "current", "", "LRD million")
    if "gdp at kp" in n or "gdp at constant" in n:
        return ("level", "constant", "2016", "LRD million")
    if "structure" in n:
        return ("share", "current", "", "percent")
    if "contribution" in n:                       # 'Contribution to growth' -> before growth
        return ("contribution", "constant", "2016", "percentage points")
    if "growth" in n:
        return ("growth_yoy", "constant", "2016", "percent")
    if "deflator" in n:
        return ("deflator", "not_applicable", "2016", "index")
    return None


def _year_header(ws):
    for r in range(1, 8):
        cols = {}
        for c in range(1, 30):
            v = ws.cell(r, c).value
            m = _YEAR_RE.match(str(v).strip()) if v is not None else None
            if m:
                cols[c] = int(m.group(1))
        if len(cols) >= 4:
            hdr_label = str(ws.cell(r, 2).value or "").strip().lower()
            return r, cols, hdr_label
    return None, {}, ""


def _num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _read_workbook(path: str, approach: str, rows: list):
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        spec = _sheet_spec(name)
        if spec is None:
            continue
        measure, basis, base, unit = spec
        ws = wb[name]
        hdr, year_cols, hdr_label = _year_header(ws)
        if not year_cols:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            code = ws.cell(r, 1).value
            label = ws.cell(r, 2).value
            label = "" if label is None else str(label).strip()
            # a repeated header ends the LRD block (USD conversion follows) -> stop
            if label.lower() == hdr_label and hdr_label:
                break
            if not label:
                continue
            low = label.lower()
            row_approach = "aggregate" if (
                "gdp at market" in low or "gross domestic" in low
                or "gross value added" in low or "gva basic" in low
            ) else approach
            for col, year in year_cols.items():
                v = _num(ws.cell(r, col).value)
                if v is None:
                    continue
                rows.append({
                    "approach": row_approach, "category": label,
                    "category_group": "",
                    "series_code": "" if code is None else str(code).strip(),
                    "geography": "National", "period": str(year),
                    "frequency": "annual", "price_basis": basis,
                    "seasonal_adjustment": "nsa", "measure": measure,
                    "value": v, "unit": unit, "base_period": base,
                })
    wb.close()


def parse(pdf_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    rows: list[dict] = []
    _read_workbook(pdf_path, "production", rows)          # primary = production
    for ex in (extras or []):
        _read_workbook(ex, "expenditure", rows)           # extra = expenditure
    if not rows:
        raise ValueError("no GDP rows parsed from LISGIS workbooks")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure", "price_basis"])
