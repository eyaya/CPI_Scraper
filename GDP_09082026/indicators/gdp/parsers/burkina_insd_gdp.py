"""Parser for the INSD Burkina Faso quarterly national-accounts (CNT) workbook.

Clean multi-sheet Excel, GDP by branch of activity (production approach). Each data
sheet shares the same layout: a header row with the periods ('2015' annual then
'YYYYTn' quarters from 2015T1), and data rows with col B = branch code (A, A1,
A1.1 …), col C = branch label, then one value per period column. Numbers use dot
decimals. The sheet-name prefix marks the series:
  o…  = gross (not seasonally adjusted)   d… = seasonally adjusted (CVS)
  PIB Crt = current prices   PIB VolCh = constant (chained volume, base 2015)
  Def = implicit deflator (2015=100)   Poids = share of GDP (published as a
  fraction)   tx4_… = YoY % change   tx1_… = QoQ % change

The PIB total row (label 'Produit Intérieur Brut') is the aggregate. Values are
billion XOF (FCFA). Everything as published; nothing derived.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

# sheet name -> (measure, price_basis, seasonal, unit)
_SHEETS = {
    "oPIB Crt":       ("level", "current", "nsa", "XOF billion"),
    "dPIB Crt":       ("level", "current", "saa", "XOF billion"),
    "oPIB VolCh":     ("level", "constant", "nsa", "XOF billion"),
    "dPIB VolCh":     ("level", "constant", "saa", "XOF billion"),
    "oDef":           ("deflator", "not_applicable", "nsa", "index"),
    "dDef":           ("deflator", "not_applicable", "saa", "index"),
    "oPoids":         ("share", "not_applicable", "nsa", "percent"),
    "dPoids":         ("share", "not_applicable", "saa", "percent"),
    "otx4_PIB Crt":   ("growth_yoy", "current", "nsa", "percent"),
    "otx4_PIB VolCh": ("growth_yoy", "constant", "nsa", "percent"),
    "otx1_PIB Crt ":  ("growth_qoq", "current", "nsa", "percent"),
    "otx1_PIB VolCh": ("growth_qoq", "constant", "nsa", "percent"),
}

_ANN_RE = re.compile(r"^(20\d\d)$")
_Q_RE = re.compile(r"^(20\d\d)T([1-4])$")

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def _read_sheet(ws):
    """Return parallel grids: cell text, numeric value or None, percent-formatted?"""
    text, num, pct = [], [], []
    for row in ws.iter_rows():
        tr, nr, pr = [], [], []
        for cell in row:
            v = cell.value
            tr.append("" if v is None else str(v).strip())
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                nr.append(None); pr.append(False)
            else:
                nr.append(float(v)); pr.append("%" in (cell.number_format or ""))
        text.append(tr); num.append(nr); pct.append(pr)
    return text, num, pct


def _period_map(text):
    """Find the header row with period labels; return (hdr_index, {col:(period,freq)})."""
    for i in range(min(8, len(text))):
        pmap = {}
        for c, t in enumerate(text[i]):
            if _Q_RE.match(t):
                m = _Q_RE.match(t); pmap[c] = (f"{m.group(1)}-Q{m.group(2)}", "quarterly")
            elif _ANN_RE.match(t):
                pmap[c] = (t, "annual")
        if len(pmap) >= 6:
            return i, pmap
    return None, {}


_CODE_RE = re.compile(r"[A-Z]+\d*(?:\.\d+)*")


def _code_label(text, r, first_pcol):
    """Read the branch code and label from the leading columns (their exact
    positions differ between sheets: level sheets have one fewer lead column
    than some growth sheets)."""
    code, label = "", ""
    for c in range(min(first_pcol, len(text[r]))):
        v = text[r][c]
        if not v:
            continue
        if len(v) <= 6 and _CODE_RE.fullmatch(v) and not code:
            code = v
        elif re.search(r"[A-Za-zÀ-ÿ]{3,}", v) and len(v) > len(label):
            label = v
    return code, label


def _parse_sheet(text, num, pct, measure, basis, seasonal, unit):
    hdr, pmap = _period_map(text)
    if not pmap:
        return []
    first_pcol = min(pmap)
    base = ""
    if basis == "constant":
        base = "chained volume, base 2015"
    elif measure == "deflator":
        base = "2015=100"
    rows = []
    for r in range(hdr + 1, len(text)):
        code, label = _code_label(text, r, first_pcol)
        if not label or len(label) < 3 or label.lower() in ("branche", "code"):
            continue
        approach = "aggregate" if "produit intérieur brut" in label.lower() else "production"
        for c, (period, freq) in pmap.items():
            if c >= len(num[r]) or num[r][c] is None:
                continue
            v = num[r][c] * 100 if pct[r][c] else num[r][c]   # %-formatted → displayed %
            rows.append({
                "approach": approach, "category": label, "category_group": "",
                "series_code": code, "geography": "National", "period": period,
                "frequency": freq, "price_basis": basis,
                "seasonal_adjustment": seasonal if freq == "quarterly" else "not_applicable",
                "measure": measure, "value": float(v), "unit": unit,
                "base_period": base,
            })
    return rows


def parse(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []
    for sheet, spec in _SHEETS.items():
        if sheet in wb.sheetnames:
            text, num, pct = _read_sheet(wb[sheet])
            rows.extend(_parse_sheet(text, num, pct, *spec))
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from INSD Burkina workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "series_code", "category", "period", "measure",
         "price_basis", "seasonal_adjustment"])
