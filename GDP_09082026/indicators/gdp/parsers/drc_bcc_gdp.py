"""Parser for the BCC (Banque Centrale du Congo) annual GDP series (xlsx).

The BCC real-sector page publishes one long aggregate GDP table, 'PIB RDC
1959-2020'. It is a compact time series (years down the rows) with columns:
  PIB constant (millions CDF, base 2000) | PIB constant (millions USD 2000) |
  taux de croissance (%) | PIB courant (millions CDF) | PIB courant (millions
  USD) | Population (millions) | PIB/habitant (USD) | PIB reel/habitant (CDF 2000).

INS-RDC (the NSO) is WAF-blocked, so BCC is the accessible source (as for CPI).
We keep the NATIVE CDF-denominated, BCC-published series — real GDP level
(constant, base 2000), nominal GDP level (current), real GDP growth, and real
GDP per capita (base 2000) — and skip the USD columns (an FX conversion, like the
Liberia USD block) and the population column (out of GDP scope). Each cell is
emitted only where BCC filled it in (the constant/growth columns run to the
2000s, the nominal column from the late 1990s), so year coverage differs by
measure. Nothing is derived or converted. Verified: growth 1960 = 1.1% equals
510,056.6 / 504,507.0 - 1.
"""
from __future__ import annotations
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def _col_spec(header: str):
    """(measure, price_basis, unit, base) for a column, or None to skip."""
    h = header.lower()
    if "population" in h:
        return None
    if "usd" in h or "dollar" in h:                # FX conversion -> skip
        return None
    if "habitant" in h:                            # PIB reel/habitant (CDF, 2000)
        return ("per_capita", "constant", "CDF", "2000")
    if "croissance" in h:
        return ("growth_yoy", "constant", "percent", "")
    if "constant" in h:
        return ("level", "constant", "million CDF", "2000")
    if "courant" in h:
        return ("level", "current", "million CDF", "")
    return None


def parse(pdf_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(pdf_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # header = row 1 (name) + row 2 (unit); classify each column
    specs = {}
    for c in range(2, ws.max_column + 1):
        header = f"{ws.cell(1, c).value or ''} {ws.cell(2, c).value or ''}"
        spec = _col_spec(header)
        if spec:
            specs[c] = spec
    rows = []
    for r in range(3, ws.max_row + 1):
        yr = ws.cell(r, 1).value                   # OBS: int on recent rows, str on old ones
        try:
            y = int(float(str(yr).strip()))
        except (ValueError, TypeError):
            continue
        if not 1900 <= y <= 2100:
            continue
        year = str(y)
        for c, (measure, basis, unit, base) in specs.items():
            v = ws.cell(r, c).value
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if abs(v) < 1e-6:                      # 0 / float-noise placeholder for a missing year
                continue
            rows.append({
                "approach": "aggregate",
                "category": "Gross Domestic Product",
                "category_group": "", "series_code": "",
                "geography": "National", "period": year,
                "frequency": "annual", "price_basis": basis,
                "seasonal_adjustment": "nsa", "measure": measure,
                "value": float(v), "unit": unit, "base_period": base,
            })
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from BCC PIB RDC workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["period", "measure", "price_basis"])
