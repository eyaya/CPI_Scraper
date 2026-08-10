"""Parser for the ANSADE Mauritania GDP workbooks (xlsx, 1998-2022).

ANSADE publishes the national accounts as a numbered set of one-sheet workbooks
(million MRU, the post-2018 new ouguiya throughout):
  1. PIB courant par branche d'activite        -> production, current level
  2. Croissance du PIB reel par branche         -> production, real growth
  3. Contribution a la croissance du PIB reel   -> production, contribution
  4. Emplois du PIB aux prix courants           -> expenditure, current level
  5. Croissance reelle des emplois              -> expenditure, real growth
We read them together (file 1 primary, the rest as extras). Each sheet has the
label in column 1 and the years (1998..2022) across the remaining columns.
Measure/approach are taken from the sheet title/name: 'emplois' -> expenditure
else production; 'croissance' -> real growth; 'contribution' -> contribution;
otherwise a current-price level. Growth and contribution cells are percent-
formatted fractions (x100 to the shown value). Verified: PIB 1998 = 38,304.8
million MRU (~US$1.0bn). Units million MRU (levels), percent (growth /
contribution). Nothing derived or converted.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

_CODE_RE = re.compile(r"^(\d+(?:\.\s*\d+)*)\s+")


def _classify(title: str):
    t = title.lower()
    approach = "expenditure" if "emploi" in t else "production"
    if "contribution" in t:
        return approach, "contribution", "constant", "percentage points"
    if "croissance" in t or "croiss" in t:
        return approach, "growth_yoy", "constant", "percent"
    return approach, "level", "current", "million MRU"        # …courant…


def _year_header(ws):
    for r in range(1, 8):
        cols = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, int)
                and 1990 <= ws.cell(r, c).value <= 2035}
        if len(cols) >= 5:
            return r, cols
    return None, {}


def _read_workbook(path, rows):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # title = the 'Tableau N: …' caption near the top, else the sheet name
    title = wb.sheetnames[0]
    for r in range(1, 5):
        for c in range(1, 4):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "tableau" in v.lower():
                title = v
                break
    approach, measure, basis, unit = _classify(title)
    is_pct = measure in ("growth_yoy", "contribution")
    hdr, year_cols = _year_header(ws)
    if not year_cols:
        wb.close()
        return
    for r in range(hdr + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or not str(label).strip():
            continue
        label = re.sub(r"\s+", " ", str(label)).strip()
        code = ""
        m = _CODE_RE.match(label)
        if m:
            code = re.sub(r"\s+", "", m.group(1))
            label = label[m.end():].strip()
        if not label:
            continue
        low = label.lower()
        row_approach = "aggregate" if (low == "pib" or "produit int" in low
                                       or low.startswith("total")) else approach
        for c, year in year_cols.items():
            cell = ws.cell(r, c)
            v = cell.value
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if is_pct and "%" in (cell.number_format or ""):
                v = v * 100.0
            rows.append({
                "approach": row_approach, "category": label, "category_group": "",
                "series_code": code, "geography": "National", "period": str(year),
                "frequency": "annual", "price_basis": basis,
                "seasonal_adjustment": "nsa", "measure": measure,
                "value": float(v), "unit": unit, "base_period": "",
            })
    wb.close()


def parse(pdf_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    rows = []
    _read_workbook(pdf_path, rows)
    for ex in (extras or []):
        _read_workbook(ex, rows)
    if not rows:
        raise ValueError("no GDP rows parsed from ANSADE Mauritania workbooks")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure", "price_basis"])
