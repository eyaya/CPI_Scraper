"""Parser for the GBoS (Gambia) annual GDP workbooks (xlsx, base year 2013).

Two published workbooks are read together (production = primary, expenditure =
extra): each has five tables as sheets — Current & Constant levels, Contribution,
Growth, Deflator. Layout is uniform: a header row carries the years across the
columns (some provisional years are suffixed '*', e.g. '2015*'); each data row
has the ISIC/line code in column 1, the industry/component label in column 2.

Scales & formats (kept exactly as published, nothing derived):
  * Production levels are in UNSCALED dalasi (GDP 2013 = 49.46bn); expenditure
    levels are in D'000 (thousand dalasi) — so units differ per workbook and are
    labelled accordingly rather than rescaled.
  * Growth & contribution cells are percent-FORMATTED fractions (0.0304 shows as
    '3.0%'); we multiply the %-formatted cells by 100 to record the displayed
    value (unit percent / percentage points).
  * Deflator is an index, base 2013 = 100.
Verified: GVA(basic) + taxes-less-subsidies = GDP at market price; production and
expenditure GDP agree once the ×1000 unit difference is accounted for.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

_YEAR_RE = re.compile(r"^(20\d\d)\*?$")


def _sheet_spec(name: str):
    """(measure, price_basis, base_period, is_pct) from the sheet name, or None."""
    n = name.lower()
    if "current" in n:
        return ("level", "current", "", False)
    if "constant" in n:
        return ("level", "constant", "2013", False)
    if "contribution" in n:
        # GBoS labels this 'Contribution to GDP' but the cells are each
        # sector/component's SHARE of GDP in percent (they sum to 100 — for
        # expenditure, Domestic Demand ~128% + Net exports ~-28% = 100, i.e. the
        # trade deficit), stored as the shown percent (25.6 = 25.6%), NOT a
        # fraction. So it is a share, recorded as-is (no x100).
        return ("share", "current", "", False)
    if "growth" in n:
        return ("growth_yoy", "constant", "2013", True)
    if "deflator" in n:
        return ("deflator", "not_applicable", "2013", False)
    return None


def _year_header(ws):
    """Return (header_row_index, {col: year_int}) scanning the first rows/cols."""
    for r in range(1, 9):
        cols = {}
        for c in range(1, 41):
            v = ws.cell(r, c).value
            m = _YEAR_RE.match(str(v).strip()) if v is not None else None
            if m:
                cols[c] = int(m.group(1))
        if len(cols) >= 4:
            return r, cols
    return None, {}


def _num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _read_workbook(path: str, approach: str, unit: str, rows: list):
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        spec = _sheet_spec(name)
        if spec is None:
            continue                      # Cover / Content
        measure, basis, base, is_pct = spec
        ws = wb[name]
        hdr, year_cols = _year_header(ws)
        if not year_cols:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            code = ws.cell(r, 1).value
            label = ws.cell(r, 2).value
            if label is None or not str(label).strip():
                continue
            label = str(label).strip()
            low = label.lower()
            row_approach = "aggregate" if (
                "gross domestic product" in low
                or "gross value added" in low
                or low.startswith("gdp")
            ) else approach
            for col, year in year_cols.items():
                cell = ws.cell(r, col)
                v = _num(cell.value)
                if v is None:
                    continue
                if is_pct and "%" in (cell.number_format or ""):
                    v *= 100.0
                if measure == "level":
                    u, unit_out = unit, unit
                elif measure == "deflator":
                    unit_out = "index"
                elif measure == "contribution":
                    unit_out = "percentage points"
                else:
                    unit_out = "percent"
                rows.append({
                    "approach": row_approach, "category": label,
                    "category_group": "",
                    "series_code": "" if code is None else str(code).strip(),
                    "geography": "National", "period": str(year),
                    "frequency": "annual", "price_basis": basis,
                    "seasonal_adjustment": "nsa", "measure": measure,
                    "value": v, "unit": unit_out, "base_period": base,
                })
    wb.close()


def parse(pdf_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    rows: list[dict] = []
    # primary = production workbook (unscaled dalasi)
    _read_workbook(pdf_path, "production", "GMD", rows)
    # extra = expenditure workbook (D'000 = thousand dalasi)
    for ex in (extras or []):
        _read_workbook(ex, "expenditure", "GMD thousand", rows)
    if not rows:
        raise ValueError("no GDP rows parsed from GBoS workbooks")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure", "price_basis"])
