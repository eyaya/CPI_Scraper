"""Parser for the INE São Tomé e Príncipe GDP workbooks (xlsx, Portuguese).

Two published workbooks are read together (production = primary, expenditure =
extra). Each is ONE sheet of vertically-stacked blocks; every block starts with a
year-header row (2008..2023) whose column-2 cell is the block title, then data
rows (col2 = label, values across the year columns). Units are 'mil STD'
(thousand new dobras); the dobra redenomination predates the series so there is
no scale break.

Block titles are ambiguous (two different blocks are both titled 'PIB A PREÇO
CORRENTE' — one is the level, one is the % structure), so the measure is decided
from the MAGNITUDE of the block's PIB/aggregate row, not the title:
  * aggregate ~millions  -> level   (unit mil STD)   basis from title (cadeia/
                                     constante/volume -> constant, else current)
  * aggregate ~100       -> share   (unit percent)   basis as above
  * aggregate small (<40)-> a growth block: kept ONLY when the title says real /
                            volume (real YoY growth); the price/nominal-growth and
                            chained-contribution blocks are DEFERRED (their
                            nominal-vs-deflator / share-vs-contribution meaning is
                            not unambiguous), rather than risk a wrong label.
Verified: current PIB 2008 = 2,763,009 and 2023 = 15,499,918 mil STD (~US$0.7bn),
VA + taxes = PIB, and by-sector VABs sum to the total. Nothing derived.
"""
from __future__ import annotations
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def _is_year_row(ws, r):
    return sum(1 for c in range(1, ws.max_column + 1)
               if isinstance(ws.cell(r, c).value, int)
               and 2000 <= ws.cell(r, c).value <= 2030) >= 8


def _year_cols(ws, r):
    return {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
            if isinstance(ws.cell(r, c).value, int)
            and 2000 <= ws.cell(r, c).value <= 2030}


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _is_aggregate(label: str) -> bool:
    low = label.lower()
    return ("pib" in low or "oferta" in low or "demanda" in low)


def _basis(title: str) -> str:
    t = title.lower()
    return "constant" if ("cadeia" in t or "constante" in t or "volume" in t) else "current"


def _classify(title: str, headline: float):
    """Return (measure, price_basis, unit) or None to defer the block."""
    t = title.lower()
    basis = _basis(title)
    if headline is None:
        return None
    a = abs(headline)
    if a > 10000:                                  # levels (millions of mil STD)
        return ("level", basis, "mil STD")
    if 40 < a <= 10000:                            # structure: PIB share == 100
        return ("share", basis, "percent")
    # small aggregate -> a growth-type block; keep only the real/volume growth
    if "volume" in t or "real" in t:
        return ("growth_yoy", "constant", "percent")
    return None                                    # price/nominal growth, contribution -> defer


def _read_workbook(path: str, approach: str, rows: list):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    r = 1
    while r <= ws.max_row:
        if not _is_year_row(ws, r):
            r += 1
            continue
        year_cols = _year_cols(ws, r)
        title = str(ws.cell(r, 2).value or "").strip()
        # the aggregate tables put a generic 'Designação' column header on the
        # year row; their real title (…Cadeia… / …volume %…) is the row above.
        if not title or "designa" in title.lower():
            for rr in (r - 1, r - 2, r - 3):
                above = str(ws.cell(rr, 2).value or "").strip() if rr >= 1 else ""
                if above and "designa" not in above.lower():
                    title = above
                    break
        # collect the block's data rows (until the next year-header)
        block = []
        rr = r + 1
        while rr <= ws.max_row and not _is_year_row(ws, rr):
            lbl = ws.cell(rr, 2).value
            if lbl is not None and str(lbl).strip():
                block.append((rr, str(lbl).strip()))
            rr += 1
        # headline = first aggregate (PIB/óptica) row's first-year value
        headline = None
        for rn, lbl in block:
            if _is_aggregate(lbl):
                for c in year_cols:
                    headline = _num(ws.cell(rn, c).value)
                    if headline:
                        break
                break
        spec = _classify(title, headline)
        if spec is not None:
            measure, basis, unit = spec
            base = "N-1" if (measure != "level" and basis == "constant") else ""
            for rn, lbl in block:
                row_approach = "aggregate" if _is_aggregate(lbl) else approach
                for c, year in year_cols.items():
                    v = _num(ws.cell(rn, c).value)
                    if v is None:
                        continue
                    rows.append({
                        "approach": row_approach, "category": lbl,
                        "category_group": "", "series_code": "",
                        "geography": "National", "period": str(year),
                        "frequency": "annual", "price_basis": basis,
                        "seasonal_adjustment": "nsa", "measure": measure,
                        "value": v, "unit": unit, "base_period": base,
                    })
        r = rr
    wb.close()


def parse(pdf_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    rows: list[dict] = []
    _read_workbook(pdf_path, "production", rows)
    for ex in (extras or []):
        _read_workbook(ex, "expenditure", rows)
    if not rows:
        raise ValueError("no GDP rows parsed from INE STP workbooks")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "period", "measure", "price_basis"])
