"""Parser for the INSTAT Madagascar rebased national accounts (TBE xlsx).

INSTAT's monthly 'Tableau de Bord de l'Economie' workbook carries the rebased
national accounts. We read the production value-added-by-branch sheets, each a
stack of captioned tables (branches in rows: code in col1, label in col2; periods
across the columns):

  Sortie_CN2   Tableau 5  VAB par branche aux prix constants (annual)  -> level, constant
               Tableau 6  VAB par branche aux prix courants  (annual)  -> level, current
  Sortie_CNT3  Tableau 11 VAB trimestrielle aux prix constants          -> level, constant, quarterly

The by-branch growth tables (Tableau 7 annual / 12 quarterly) are DEFERRED: their
cells mix scales (some fractions, some already percent) and this broken workbook
does not expose a reliable number-format to disambiguate, so recording them would
risk a wrong x100.
Units billion Ariary (levels). The base year of the constant series is not stated
in the workbook, so base_period is left blank rather than guessed. The aggregate
'COMPTES NATIONAUX' summary sheet is formula-driven and reads as #REF! in values
mode, so it is skipped; the expenditure 'origines et emplois' tables and the
inflation table are left for a later pass. Everything as published; nothing
derived. (The file has broken chart drawings, so it must be opened read-only.)
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

_SHEETS = ("sortie_cn2", "sortie_cnt3")
_QUARTER_RE = re.compile(r"^(20\d\d)[:\-\s]?Q([1-4])$", re.I)


def _spec(caption: str):
    c = caption.lower()
    if "valeur ajout" in c and "constant" in c:
        return ("level", "constant", "billion Ariary")
    if "valeur ajout" in c and "courant" in c:
        return ("level", "current", "billion Ariary")
    # the growth tables (Tableau 7/12) store values at inconsistent scales
    # (some as fractions, some already as percent) with no reliable number-format
    # to tell them apart in this broken workbook, so they are deferred rather than
    # risk a wrong x100.
    return None


def _period(v):
    if isinstance(v, int) and 1990 <= v <= 2035:
        return str(v)
    s = str(v).strip()
    m = _QUARTER_RE.match(s)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"
    if re.fullmatch(r"20\d\d", s):
        return s
    return None


def _year_header(grid, r0, r1):
    for r in range(r0, min(r1, len(grid))):
        cols = {c: _period(grid[r][c]) for c in range(len(grid[r]))
                if _period(grid[r][c])}
        if len(cols) >= 3:
            return r, cols
    return None, {}


def _num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def parse(pdf_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(pdf_path, data_only=True, read_only=True)
    rows = []
    for name in wb.sheetnames:
        if name.lower() not in _SHEETS:
            continue
        ws = wb[name]
        grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        n = len(grid)
        # locate caption rows ('Tableau N: ...')
        caps = [(i, str(grid[i][0])) for i in range(n)
                if grid[i] and isinstance(grid[i][0], str)
                and re.match(r"\s*tableau\s*\d", grid[i][0], re.I)]
        for k, (ci, caption) in enumerate(caps):
            spec = _spec(caption)
            if spec is None:
                continue
            measure, basis, unit = spec
            end = caps[k + 1][0] if k + 1 < len(caps) else n
            hdr, year_cols = _year_header(grid, ci + 1, min(ci + 6, end))
            if not year_cols:
                continue
            freq = "quarterly" if any("Q" in p for p in year_cols.values()) else "annual"
            for r in range(hdr, end):
                code = grid[r][0] if len(grid[r]) > 0 else None
                label = grid[r][1] if len(grid[r]) > 1 else None
                if label is None or not str(label).strip():
                    continue
                label = re.sub(r"\s+", " ", str(label)).strip()
                if re.match(r"tableau\s*\d", label, re.I) or label.lower() == "branches d'activité":
                    continue
                low = label.lower()
                approach = "aggregate" if (low == "pib" or "produit int" in low) else "production"
                for c, period in year_cols.items():
                    v = _num(grid[r][c] if c < len(grid[r]) else None)
                    if v is None:
                        continue
                    rows.append({
                        "approach": approach, "category": label, "category_group": "",
                        "series_code": "" if code is None else str(code).strip(),
                        "geography": "National", "period": period,
                        "frequency": freq, "price_basis": basis,
                        "seasonal_adjustment": "nsa", "measure": measure,
                        "value": v, "unit": unit, "base_period": "",
                    })
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from INSTAT Madagascar TBE workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure", "price_basis"])
