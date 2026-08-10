"""Parser for the Rwanda NISR 'GDP National Accounts' quarterly workbook.

Wide time series; column layout on every data sheet: col B = description, col C =
ISIC4 code (production sheets), col D = latest annual figure (2024), col E onward =
quarters from 2006 Q1. The header row is located, not hardcoded (it varies).

Production sheets are one measure each; the two expenditure sheets stack several
measures in labelled sub-blocks:
  QGDP_CP   production level current            QGDP_Cont production contribution
  QGDP_KP   production level constant 2024      QGDP_DF   production deflator 2024=100
  QGDP_Gr   production real growth              QGDP_SH   production share of GDP
  T3 GDP XQ  expenditure: 'In current prices' / 'In constant prices' / 'Deflators'
  T3A GDP XQ expenditure: 'Percentage of GDP…' / 'Growth rates…'

NISR percent-formats its ratio cells (growth & shares stored as fractions, e.g.
0.10 displayed as "10%"). We read each cell's number format and scale %-formatted
cells to the published percentage — i.e. we record exactly what the workbook
displays. Nothing else is derived. The 'Table A Q' summary sheet is skipped.
"""
from __future__ import annotations
import re
import openpyxl

_QUARTER_RE = re.compile(r"^(\d{4})\s*Q\s*([1-4])$")
_YEAR_RE = re.compile(r"^(\d{4})(?:\.0)?$")

# single-measure production sheets: sheet -> (measure, price_basis, unit)
_PROD_SHEETS = {
    "QGDP_CP":   ("level",        "current",        "RWF billion"),
    "QGDP_KP":   ("level",        "constant",       "RWF billion"),
    "QGDP_Gr":   ("growth_yoy",   "constant",       "percent"),
    "QGDP_Cont": ("contribution", "constant",       "percentage points"),
    "QGDP_DF":   ("deflator",     "not_applicable", "index"),
    "QGDP_SH":   ("share",        "not_applicable", "percent"),
}
_EXP_SHEETS = ["T3 GDP XQ", "T3A GDP XQ"]

# expenditure sub-block header (lowercased, priority order) -> (measure, basis, unit)
_EXP_BLOCKS = [
    ("percentage of gdp", ("share", "not_applicable", "percent")),
    ("growth rate",       ("growth_yoy", "constant", "percent")),
    ("deflator",          ("deflator", "not_applicable", "index")),
    ("constant price",    ("level", "constant", "RWF billion")),
    ("current price",     ("level", "current", "RWF billion")),
]

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def _read_sheet(ws):
    """Return (text, num, pct) parallel grids: cell text, numeric value or None,
    and whether the cell is percent-formatted."""
    text, num, pct = [], [], []
    for row in ws.iter_rows():
        tr, nr, pr = [], [], []
        for cell in row:
            v = cell.value
            tr.append("" if v is None else str(v).strip())
            if isinstance(v, bool):
                nr.append(None); pr.append(False)
            elif isinstance(v, (int, float)):
                nr.append(float(v)); pr.append("%" in (cell.number_format or ""))
            else:
                nr.append(None); pr.append(False)
        text.append(tr); num.append(nr); pct.append(pr)
    return text, num, pct


def _find_header(text):
    best_row, best_map = None, {}
    for r in range(min(10, len(text))):
        pmap = {}
        for c in range(2, len(text[r])):
            t = text[r][c]
            qm = _QUARTER_RE.match(t)
            if qm:
                pmap[c] = (f"{qm.group(1)}-Q{qm.group(2)}", "quarterly")
            elif _YEAR_RE.match(t):
                pmap[c] = (_YEAR_RE.match(t).group(1), "annual")
        if len(pmap) > len(best_map):
            best_row, best_map = r, pmap
    if not best_map:
        raise ValueError("no period header row found")
    return best_row, best_map


def _base_period(measure, basis, base_year):
    if measure == "deflator":
        return f"{base_year}=100" if base_year else "index"
    if basis == "constant":
        return f"constant {base_year} prices" if base_year else "constant prices"
    return ""


def _emit(text, num, pct, r, pmap, approach, measure, basis, unit, base_year,
          isic="", label=None):
    if label is None:
        label = re.sub(r"\s+", " ", text[r][1]).strip()
    if not label:
        return []
    row_approach = "aggregate" if label.lower().startswith("gross domestic product") else approach
    base = _base_period(measure, basis, base_year)
    seasonal_q = "nsa"
    out = []
    for c, (period, freq) in pmap.items():
        if c >= len(num[r]) or num[r][c] is None:
            continue
        val = num[r][c] * 100 if pct[r][c] else num[r][c]
        out.append({
            "approach": row_approach, "category": label, "category_group": "",
            "series_code": isic, "geography": "National", "period": period,
            "frequency": freq, "price_basis": basis,
            "seasonal_adjustment": seasonal_q if freq == "quarterly" else "not_applicable",
            "measure": measure, "value": float(val), "unit": unit, "base_period": base,
        })
    return out


def _base_year_of(text):
    joined = " ".join(text[r][1] for r in range(min(6, len(text))) if len(text[r]) > 1)
    m = re.search(r"constant\s+(\d{4})", joined, re.I)
    return m.group(1) if m else "2024"


def _parse_production(text, num, pct, measure, basis, unit):
    hdr, pmap = _find_header(text)
    by = _base_year_of(text)
    rows = []
    for r in range(hdr + 1, len(text)):
        isic = text[r][2] if len(text[r]) > 2 else ""
        rows.extend(_emit(text, num, pct, r, pmap, "production", measure, basis,
                          unit, by, isic=isic))
    return rows


def _match_block(label):
    l = label.lower()
    for key, spec in _EXP_BLOCKS:
        if key in l:
            return spec
    return None


def _parse_expenditure(text, num, pct):
    hdr, pmap = _find_header(text)
    rows, block, section = [], None, ""
    for r in range(hdr + 1, len(text)):
        label = re.sub(r"\s+", " ", text[r][1]).strip() if len(text[r]) > 1 else ""
        if not label:
            continue
        has_data = any(c < len(num[r]) and num[r][c] is not None for c in pmap)
        low = label.lower()
        if "export" in low:
            section = "Exports"
        elif "import" in low:
            section = "Imports"
        blk = _match_block(label)
        if blk and not has_data:            # a sub-block header row
            block = blk
            continue
        if block is None or not has_data:
            continue
        # 'Goods (fob)' and 'Services' repeat under Exports and Imports; keep the
        # two series distinct by tagging them with the section.
        disp = f"{label} ({section})" if section and low in ("goods (fob)", "services") else label
        measure, basis, unit = block
        rows.extend(_emit(text, num, pct, r, pmap, "expenditure", measure, basis,
                          unit, "2024", label=disp))
    return rows


def parse(xlsx_path: str):
    import pandas as pd
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []
    for sheet, (measure, basis, unit) in _PROD_SHEETS.items():
        if sheet in wb.sheetnames:
            text, num, pct = _read_sheet(wb[sheet])
            rows += _parse_production(text, num, pct, measure, basis, unit)
    for sheet in _EXP_SHEETS:
        if sheet in wb.sheetnames:
            text, num, pct = _read_sheet(wb[sheet])
            rows += _parse_expenditure(text, num, pct)
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from Rwanda NISR workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS]
