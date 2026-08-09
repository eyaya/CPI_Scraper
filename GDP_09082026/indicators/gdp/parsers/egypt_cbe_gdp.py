"""Parser for the Central Bank of Egypt GDP time series (xlsx).

CBE publishes four GDP workbooks — GDP at factor cost and GDP by expenditure, each
at constant (base 2006/2007) and current prices. Each workbook has ONE SHEET PER
FISCAL YEAR (Egypt's fiscal year runs July-June), and every sheet holds quarterly
figures split into Public / Private / Total columns per quarter:

  row 'Q1 … Q2 … Q3 … Q4'   then   'Public Private Total' × 4
  then one row per sector/component (col A English label, col B Arabic).

We read the TOTAL column of each quarter (the Public/Private split is left for a
later pass) and map the fiscal quarter to its calendar quarter — FY Q1 = Jul-Sep,
Q2 = Oct-Dec, Q3 = Jan-Mar, Q4 = Apr-Jun — so fiscal year 2022/2023 Q1 becomes
2022-Q3, Q3 becomes 2023-Q1, etc. Units differ by workbook: factor-cost is EGP
million, expenditure EGP billion (they reconcile). The constant base year is read
from each sheet's 'at YYYY/YYYY prices' caption (Egypt reprices to the previous
year / rebases periodically). factor-cost workbooks are the
production approach, expenditure workbooks the expenditure approach. Verified: the
'GDP' row totals to the published quarterly GDP. Nothing derived (the fiscal->
calendar quarter is a faithful relabel of the same three months).
"""
from __future__ import annotations
import os
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

# fiscal-quarter index (0..3, in Q1..Q4 order) -> (calendar-year offset, calendar Q)
_FQ_TO_CAL = {0: (0, 3), 1: (0, 4), 2: (1, 1), 3: (1, 2)}


def _spec(filename: str):
    n = filename.lower()
    if "expenditure" in n:
        return "expenditure", ("constant" if "constant" in n else "current"), "EGP billion"
    return "production", ("constant" if "constant" in n else "current"), "EGP million"


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _value_cols(grid):
    """Return (data_start_row, [4 value columns]) for whichever layout the sheet
    uses: factor-cost has Public/Private/Total per quarter (take the 4 Total cols);
    expenditure has one column per quarter (take the Q1..Q4 columns)."""
    for i, row in enumerate(grid[:12]):
        qcols = {str(v).strip().upper(): c for c, v in enumerate(row)
                 if isinstance(v, str) and re.fullmatch(r"Q[1-4]", str(v).strip(), re.I)}
        if len(qcols) >= 4:
            nxt = grid[i + 1] if i + 1 < len(grid) else []
            totals = [c for c, v in enumerate(nxt)
                      if isinstance(v, str) and v.strip().lower() == "total"]
            if len(totals) >= 4:
                return i + 2, totals[:4]
            return i + 1, [qcols[f"Q{q}"] for q in range(1, 5)]
    return None, []


def _read_workbook(path, rows):
    approach, basis, unit = _spec(os.path.basename(path))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        m = re.match(r"(\d{4})[-/](\d{4})", sheet.strip())
        if not m:
            continue
        fy = int(m.group(1))
        ws = wb[sheet]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        data_start, value_cols = _value_cols(grid)
        if not value_cols:
            continue
        # base year: the sheet states 'at YYYY/YYYY prices' (Egypt uses the previous
        # year's prices); blank for the current-price workbooks.
        base = ""
        if basis == "constant":
            for row in grid[:data_start]:
                for v in row:
                    b = re.search(r"at\s*(\d{4}\s*/\s*\d{4})\s*prices",
                                  str(v), re.I) if isinstance(v, str) else None
                    if b:
                        base = re.sub(r"\s+", "", b.group(1))
                        break
                if base:
                    break
        for r in range(data_start, len(grid)):
            row = grid[r]
            label = row[0] if row else None
            if not (isinstance(label, str) and re.search(r"[A-Za-z]{3}", label)):
                continue
            label = re.sub(r"\s+", " ", label).strip(" .:-")
            low = label.lower()
            row_approach = "aggregate" if ("gross domestic" in low or "gdp" in low
                                           or "domestic product" in low) else approach
            for qi, c in enumerate(value_cols):
                v = _num(row[c] if c < len(row) else None)
                if v is None:
                    continue
                off, cq = _FQ_TO_CAL[qi]
                rows.append({
                    "approach": row_approach, "category": label,
                    "category_group": "", "series_code": "",
                    "geography": "National", "period": f"{fy + off}-Q{cq}",
                    "frequency": "quarterly", "price_basis": basis,
                    "seasonal_adjustment": "nsa", "measure": "level",
                    "value": v, "unit": unit, "base_period": base,
                })
    wb.close()


def parse(pdf_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    rows = []
    _read_workbook(pdf_path, rows)
    for ex in (extras or []):
        _read_workbook(ex, rows)
    if not rows:
        raise ValueError("no GDP rows parsed from CBE Egypt workbooks")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "period", "measure", "price_basis"])
