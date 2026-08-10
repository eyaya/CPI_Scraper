"""Parser for the ANSD Senegal renovated national-accounts workbook (xlsx, base 2021).

'Tableaux-publication...Base2021.xlsx'. We read three of its sheets (years across
the columns, 2021-2024, million FCFA / XOF):

  'Les trois approches du PIB'   -> GDP by the production, expenditure AND income
      approaches (current prices), each block opened by a 'B.1 Produit interieur
      brut' aggregate row, with SNA operation codes (B.1b, D.21, P.3, P.5, D.1 ...).
  'PIB par secteur d'activite'   -> GDP/value added by ISIC activity (production,
      current prices), with primary/secondary/tertiary sub-totals.
  'Agregats'                     -> the constant-price PIB series and the real GDP
      growth rate (the current-price aggregates there duplicate the approaches
      sheet and are skipped; per-capita, incl. a USD FX column, is left for later).

Layout is detected dynamically: the year-header row carries 2021..2024; a data
row's label is the right-most text cell before the first year column and its code
(SNA/ISIC) the cell before that. The growth cells are percent-formatted fractions
(x100 to the shown value). Verified: PIB market price 2024 = 22,745,886 = VA
21,097,520 + taxes 2,234,034 - subsidies 585,668; the three approaches all equal
PIB each year. Units million FCFA (levels), percent (growth). Nothing derived.
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]

_APPROACH_BLOCKS = ["production", "expenditure", "income"]
_CODE_RE = re.compile(r"^[A-Z]{1,2}[\.\d][\.\d]*[a-z]?$|^[A-Z]\d{2}$|^[A-Z]{2}$")


def _norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def _is_pib(label, code):
    low = label.lower()
    return code == "B.1" or low.startswith("produit int") or low == "pib" \
        or "pib au prix du march" in low


def _year_header(ws):
    for r in range(1, 12):
        cols = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, int)
                and 2015 <= ws.cell(r, c).value <= 2035}
        if len(cols) >= 3:
            return r, cols
    return None, {}


def _label_and_code(ws, r, first_year_col):
    label, code = "", ""
    for c in range(1, first_year_col):
        v = ws.cell(r, c).value
        if v is None or not str(v).strip():
            continue
        s = _norm(v)
        if _CODE_RE.match(s) and not re.search(r"[a-z]{3}", s):
            code = s
        elif re.search(r"[A-Za-z]{3}", s):
            label = s
    return label, code


def _all_year_headers(ws):
    out = []
    for r in range(1, ws.max_row + 1):
        cols = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, int)
                and 2015 <= ws.cell(r, c).value <= 2035}
        if len(cols) >= 3:
            out.append((r, cols))
    return out


def _sector_block_spec(title: str):
    """Classify a 'PIB par secteur' sub-table from its title; None to defer."""
    t = title.lower()
    if not t:
        return None                                # untitled block (structure/contrib) -> defer
    if "volume" in t and ("%" in t or "evolution" in t or "évolution" in t):
        return ("growth_yoy", "not_applicable", "")
    if "volume" in t or "chain" in t or "chaîn" in t or "constant" in t:
        m = re.search(r"r[eé]f[eé]rence (\d{4})", t)
        return ("level", "constant", m.group(1) if m else "")
    if "courant" in t:
        return ("level", "current", "")
    return None


def _read_sector(ws, rows):
    headers = _all_year_headers(ws)
    for i, (hdr, year_cols) in enumerate(headers):
        end = headers[i + 1][0] - 1 if i + 1 < len(headers) else ws.max_row
        # title = nearest non-empty text cell in col2 above this header
        title = ""
        for rr in range(hdr - 1, max(0, hdr - 5), -1):
            v = ws.cell(rr, 2).value
            if v and len(_norm(v)) > 6:
                title = _norm(v)
                break
        spec = _sector_block_spec(title)
        if spec is None:
            continue
        measure, basis, base = spec
        fyc = min(year_cols)
        for r in range(hdr + 1, end + 1):
            label, code = _label_and_code(ws, r, fyc)
            if not label or label.lower() in ("opérations", "operations", "code"):
                continue
            approach = "aggregate" if _is_pib(label, code) else "production"
            for c, year in year_cols.items():
                cell = ws.cell(r, c)
                v = cell.value
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                if measure == "growth_yoy" and "%" in (cell.number_format or ""):
                    v = v * 100.0
                unit = "percent" if measure == "growth_yoy" else "million FCFA"
                rows.append({
                    "approach": approach, "category": label, "category_group": "",
                    "series_code": code, "geography": "National", "period": str(year),
                    "frequency": "annual", "price_basis": basis,
                    "seasonal_adjustment": "nsa", "measure": measure,
                    "value": float(v), "unit": unit, "base_period": base,
                })


def _read_sheet(ws, kind, rows):
    if kind == "sector":
        _read_sector(ws, rows)
        return
    hdr, year_cols = _year_header(ws)
    if not year_cols:
        return
    fyc = min(year_cols)
    block_i = -1                                   # for the 'three approaches' sheet
    mode = ("level", "current", "")                # for the 'agregats' sheet sections
    for r in range(hdr + 1, ws.max_row + 1):
        label, code = _label_and_code(ws, r, fyc)
        if not label:
            continue
        vals = {y: ws.cell(r, c) for c, y in year_cols.items()}
        has_num = any(isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
                      for c in vals.values())

        if kind == "agregats":
            low = label.lower()
            if not has_num:                        # section header row
                if "constant" in low:
                    # the Agregats constant block is labelled 'de 2014' but the
                    # sector 'volume' table states 'reference 2021'; take the
                    # constant series from there and skip this conflicting one.
                    mode = ("skip", "", "")
                elif "habitant" in low:
                    mode = ("skip", "", "")
                elif "croissance" in low:
                    mode = ("growth_yoy", "not_applicable", "")
                continue
            if "croissance" in low:                # real GDP growth row (has values inline)
                measure, basis, base = "growth_yoy", "not_applicable", ""
            else:
                measure, basis, base = mode
            if measure == "skip":
                continue
            # the current-price aggregates duplicate the approaches sheet
            if measure == "level" and basis == "current" and not (
                    "revenu national" in low):
                continue
            approach = "aggregate"
        elif kind == "sector":
            measure, basis, base = "level", "current", ""
            approach = "aggregate" if _is_pib(label, code) else "production"
            if not has_num:
                continue
        else:  # three approaches
            if not has_num:
                continue
            measure, basis, base = "level", "current", ""
            if _is_pib(label, code):
                block_i += 1
                approach = "aggregate"
            else:
                approach = _APPROACH_BLOCKS[min(block_i, 2)] if block_i >= 0 else "production"

        for c, year in year_cols.items():
            cell = ws.cell(r, c)
            v = cell.value
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if measure == "growth_yoy" and "%" in (cell.number_format or ""):
                v = v * 100.0
            unit = "percent" if measure == "growth_yoy" else "million FCFA"
            rows.append({
                "approach": approach, "category": label, "category_group": "",
                "series_code": code, "geography": "National", "period": str(year),
                "frequency": "annual", "price_basis": basis,
                "seasonal_adjustment": "nsa", "measure": measure,
                "value": float(v), "unit": unit, "base_period": base,
            })


def parse(pdf_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(pdf_path, data_only=True)
    rows = []
    for name in wb.sheetnames:
        low = name.lower()
        if "approches" in low:
            _read_sheet(wb[name], "approaches", rows)
        elif "secteur" in low:
            _read_sheet(wb[name], "sector", rows)
        elif low.startswith("agr"):
            _read_sheet(wb[name], "agregats", rows)
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from ANSD Senegal workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure", "price_basis"])
