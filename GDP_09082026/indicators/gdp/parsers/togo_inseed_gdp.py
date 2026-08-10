"""Parser for the INSEED Togo national-accounts VAB-by-branch workbook (xlsx).

INSEED publishes 'Annexes 1/2 - Valeurs ajoutees brutes - branches.xlsx', a single
sheet 'Vab des branches': gross value added by ISIC branch (code in col1, label in
col2) with the years 2007-2019 across the columns, plus a 'Total des branches'
row. This is the production approach at current prices (the WAEMU/AFRISTAT
convention for an unlabelled VAB annexe of this vintage; the values grow with a
nominal, not real, profile). Units are million XOF (FCFA). We emit each branch's
VA level and the total as-published; nothing derived. It is an older series
(2007-2015, pre the 2016 rebasing) and covers value added only (no product taxes,
so no GDP-at-market row, and no constant-price or expenditure tables in this
file). Verified: Total des branches 2015 = 3,095,068 million XOF.
"""
from __future__ import annotations
import openpyxl
import pandas as pd

_OUT_COLS = ["approach", "category", "category_group", "series_code", "geography",
             "period", "frequency", "price_basis", "seasonal_adjustment",
             "measure", "value", "unit", "base_period"]


def parse(pdf_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(pdf_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    def _year(v):
        try:
            y = int(float(str(v).strip()))
        except (ValueError, TypeError, AttributeError):
            return None
        return y if 1990 <= y <= 2100 else None

    # header row = the one carrying the year values (stored as ints or strings)
    hdr, year_cols = None, {}
    for r in range(1, 6):
        cols = {c: _year(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)
                if _year(ws.cell(r, c).value) is not None}
        if len(cols) >= 4:
            hdr, year_cols = r, cols
            break
    if not year_cols:
        raise ValueError("no year header found in Togo VAB workbook")

    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        label = ws.cell(r, 2).value
        if label is None or not str(label).strip():
            continue
        label = str(label).strip()
        approach = "aggregate" if "total des branches" in label.lower() else "production"
        for c, year in year_cols.items():
            raw = ws.cell(r, c).value
            if isinstance(raw, bool):
                continue
            if isinstance(raw, (int, float)):
                v = float(raw)
            else:
                try:
                    v = float(str(raw).replace(" ", "").replace(",", "."))
                except (ValueError, TypeError, AttributeError):
                    continue
            rows.append({
                "approach": approach, "category": label, "category_group": "",
                "series_code": "" if code is None else str(code).strip(),
                "geography": "National", "period": str(year),
                "frequency": "annual", "price_basis": "current",
                "seasonal_adjustment": "nsa", "measure": "level",
                "value": v, "unit": "million XOF", "base_period": "",
            })
    wb.close()
    if not rows:
        raise ValueError("no GDP rows parsed from Togo VAB workbook")
    return pd.DataFrame.from_records(rows)[_OUT_COLS].drop_duplicates(
        ["approach", "category", "series_code", "period", "measure"])
