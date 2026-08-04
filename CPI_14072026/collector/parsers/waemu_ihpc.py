"""Parser for the WAEMU/UEMOA harmonised CPI (IHPC), shared across member states.

The 8 UEMOA NSOs run the same PHÉNIX-UEMOA methodology and publish a 'série de
données' workbook: a wide monthly index time series (from 1998, spliced/
'raccordées' to base 2023 = 100) with one row per COICOP-2018 division (codes
01..13) plus an all-items 'Global'/'Ensemble' row. The exact shape varies a
little by country — Senegal (ANSD) uses a 'DIV' sheet with a 'Code' column and a
header offset; Benin (INStaD) a single sheet with a 'DIVISIONS' column at row 0
— so we detect the layout rather than hard-code it:

  * the data sheet is the first worksheet whose header row has >=12 month
    (datetime) columns and a code column headed 'Code' or 'DIVISIONS';
  * division rows carry a 2-digit code (01..13); the all-items row is the one
    labelled 'Global' or 'Ensemble' (code 00).

Index levels only (the sheets carry no rates; YoY/MoM derivable). Labels kept as
published (French). This one parser serves every UEMOA country — add a country
by pointing a descriptor at its série-de-données file.
"""
from __future__ import annotations
import datetime as dt
import re
import openpyxl
import pandas as pd

_BASE_PERIOD = "2023 = 100"
_CODE_HEADERS = {"code", "divisions", "division"}
_ALL_ITEMS = {"global", "ensemble"}


def _code(raw) -> str | None:
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)) and 1 <= raw <= 13 and float(raw).is_integer():
        return f"{int(raw):02d}"
    if isinstance(raw, str) and re.fullmatch(r"\d{1,2}", raw.strip()):
        n = int(raw.strip())
        return f"{n:02d}" if 1 <= n <= 13 else None
    return None


def _extract_sheet(ws) -> list[tuple]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows)
               if sum(isinstance(c, (dt.datetime, dt.date)) for c in r) >= 12), None)
    if hi is None:
        return []
    hdr = rows[hi]
    code_col = next((j for j, c in enumerate(hdr)
                     if isinstance(c, str) and c.strip().lower() in _CODE_HEADERS), None)
    if code_col is None:
        return []
    label_col = next((j for j, c in enumerate(hdr)
                      if isinstance(c, str) and "libell" in c.lower()), None)
    months = {j: f"{c.year}-{c.month:02d}" for j, c in enumerate(hdr)
              if isinstance(c, (dt.datetime, dt.date))}

    out = []
    for r in rows[hi + 1:]:
        code = _code(r[code_col] if code_col < len(r) else None)
        lab = r[label_col] if (label_col is not None and label_col < len(r)) else None
        if code is None:
            if isinstance(lab, str) and lab.strip().lower() in _ALL_ITEMS:
                code = "00"
            else:
                continue
        label = "All items" if code == "00" else (str(lab).strip() if lab else "")
        for j, period in months.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append((code, label, period, round(float(v), 4)))
    return out


def parse(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []
    for name in wb.sheetnames:
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):        # skip chart sheets
            continue
        recs = _extract_sheet(ws)
        if len({c for c, _, _, _ in recs if c != "00"}) >= 12:   # the divisions sheet
            records = recs
            break
    if not records:
        raise ValueError("no WAEMU IHPC division sheet found in workbook")

    out = pd.DataFrame.from_records(
        records, columns=["coicop_code", "coicop_label", "period", "value"])
    out = out.drop_duplicates(["coicop_code", "period"])
    out["geography"] = "National"
    out["measure"] = "index"
    out["unit"] = "Index"
    out["base_period"] = _BASE_PERIOD
    out["frequency"] = "monthly"
    return out
