"""Parser for the Morocco HCP IPC (Indice des Prix a la Consommation), Tier 2.

HCP publishes the CPI (base 2017 = 100) as Google Sheets it updates monthly and
links from its IPC pages. The by-division data lives in one workbook with two
'mensuel' tabs (groupe 1 = COICOP divisions 01..06, groupe 2 = 07..12), a wide
monthly time series (2016 -> latest). The overall 'Indice General' (All items,
00) sits in a separate 'grandes divisions' workbook, passed in as `extras`.

Each sheet is the same shape: a header row with 'Mois' + one column per series,
then rows of '<YYYY/MM>  <value> ...' (newest first) ending in a 'Source:'
footer. We read index levels only (the sheets carry no rates); YoY/MoM are
derivable downstream. Morocco uses COICOP-1999 (12 divisions + general = 13).
"""
from __future__ import annotations
import datetime as dt
import re
import unicodedata
import openpyxl
import pandas as pd

_BASE_PERIOD = "2017 = 100"

# code -> keyword (on the accent-stripped, lower-cased French header label).
# Order matters: check All-items first so 'indice general' isn't mis-hit.
_DIVISION_KEYS = [
    ("00", "indice general"),
    ("01", "produits alimentaires"),
    ("02", "tabac"),
    ("03", "habillement"),
    ("04", "logement"),
    ("05", "meubles"),
    ("06", "sante"),
    ("07", "transport"),
    ("08", "communication"),
    ("09", "loisirs"),
    ("10", "enseignement"),
    ("11", "restaurants"),
    ("12", "biens et services divers"),
]


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _code_for(label: str, allowed: set[str]) -> str | None:
    n = _norm(label)
    for code, kw in _DIVISION_KEYS:
        if code in allowed and kw in n:
            return code
    return None


def _period(v) -> str | None:
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.year}-{v.month:02d}"
    if isinstance(v, str):
        m = re.match(r"\s*(\d{4})[/-](\d{1,2})", v)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None


def _extract(ws, allowed: set[str]) -> list[tuple]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows)
               if any(isinstance(c, str) and c.strip().lower() == "mois" for c in r)), None)
    if hi is None:
        return []
    hdr = rows[hi]
    mois_col = next(j for j, c in enumerate(hdr)
                    if isinstance(c, str) and c.strip().lower() == "mois")
    cols = {}
    for j, c in enumerate(hdr):
        if isinstance(c, str):
            code = _code_for(c, allowed)
            if code:
                cols[j] = (code, c.strip())
    out = []
    for r in rows[hi + 1:]:
        period = _period(r[mois_col]) if mois_col < len(r) else None
        if not period:
            continue                       # skips the trailing 'Source:' footer
        for j, (code, label) in cols.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                out.append((code, label, period, round(float(v), 4)))
    return out


def parse(xlsx_path: str, extras: list[str] | None = None) -> pd.DataFrame:
    records = []

    # primary workbook: the 12 divisions across the two 'mensuel' tabs
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    div_codes = {c for c, _ in _DIVISION_KEYS if c != "00"}
    for sn in wb.sheetnames:
        if "mensuel" in sn.lower():
            records += _extract(wb[sn], div_codes)

    # extras: the 'grandes divisions' workbook holds the general index (00)
    for ex in extras or []:
        wbx = openpyxl.load_workbook(ex, data_only=True)
        for sn in wbx.sheetnames:
            if "mensuel" in sn.lower():
                records += _extract(wbx[sn], {"00"})

    if not records:
        raise ValueError("no IPC series extracted from HCP workbook(s)")

    out = pd.DataFrame.from_records(
        records, columns=["coicop_code", "coicop_label", "period", "value"])
    out = out.drop_duplicates(["coicop_code", "period"])
    out["geography"] = "National"
    out["measure"] = "index"
    out["unit"] = "Index"
    out["base_period"] = _BASE_PERIOD
    out["frequency"] = "monthly"
    return out
