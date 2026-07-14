"""Parser for the CAPMAS monthly 'Consumer Price Index' bulletin Excel (Tier 2).

CAPMAS — Egypt's NSO — publishes the CPI each month as a multi-sheet workbook,
reachable through the site's JSON API (publication id 32). This is the primary,
authoritative Egypt source; the CBE inflation-note parser (egypt_cbe) is a
Tier-3 fallback that only carries the urban series.

'Table 1' gives the CPI by COICOP-1999 division for THREE geographies side by
side — Urban, Rural and Total Egypt — each block being:

    <weights>  <index, current month>  <MoM %>  <YoY %>       (base 2018/2019 = 100)

The table interleaves the 12 divisions (+ All items) with their sub-groups. The
13 division rows carry an ALL-CAPS English label in the last column, whereas
sub-groups are Title Case, so we key on the 13 known division labels and ignore
the rest. We emit index + inflation_mom + inflation_yoy for the report month,
for each of the three geographies (39 series/month once, per geography 13).
"""
from __future__ import annotations
import re
import openpyxl
import pandas as pd

# COICOP-1999 division -> (code, canonical label, predicate on the NORMALISED
# (upper-cased, punctuation->space) English label). Egypt's wording is its own,
# so we match here rather than via the shared COICOP-2018 code_for_label().
_DIV_RULES = [
    ("00", "All items",                                                 lambda s: s == "ALL ITEMS"),
    ("01", "Food and non-alcoholic beverages",                          lambda s: s.startswith("FOOD AND NON")),
    ("02", "Alcoholic beverages, tobacco and narcotics",                lambda s: "TOBACCO" in s and "NARCOTIC" in s),
    ("03", "Clothing and footwear",                                     lambda s: s.startswith("CLOTHING AND FOOTWEAR")),
    ("04", "Housing, water, electricity, gas and other fuels",          lambda s: s.startswith("HOUSING")),
    ("05", "Furnishings, household equipment and routine maintenance",  lambda s: s.startswith("FURNISHING")),
    ("06", "Health",                                                    lambda s: s == "HEALTH"),
    ("07", "Transport",                                                 lambda s: s == "TRANSPORT"),
    ("08", "Communications",                                            lambda s: s == "COMMUNICATIONS"),
    ("09", "Recreation and culture",                                    lambda s: s.startswith("RECREATION AND CULTURE")),
    ("10", "Education",                                                 lambda s: s == "EDUCATION"),
    ("11", "Restaurants and hotels",                                    lambda s: s.startswith("RESTAURANTS AND HOTELS")),
    ("12", "Miscellaneous goods and services",                          lambda s: s.startswith("MISCELLANEOUS GOODS")),
]

# canonical geography name -> token that identifies its header cell ('X Egypt')
_GEOS = [("Urban", "URBAN"), ("Rural", "RURAL"), ("Total", "TOTAL")]
_BASE_PERIOD = "2018/2019 = 100"


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]", " ", str(s).upper())).strip()


def _table1(wb):
    for name in wb.sheetnames:
        if "table 1" in name.lower():
            return wb[name]
    raise ValueError(f"'Table 1' sheet not found in {wb.sheetnames}")


def parse(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = [list(r) for r in _table1(wb).iter_rows(values_only=True)]

    # header row = the one carrying 'weights' at the start of each geography block
    hdr_i = next(
        (i for i, r in enumerate(rows)
         if sum(isinstance(c, str) and c.strip().lower() == "weights" for c in r) >= 3),
        None,
    )
    if hdr_i is None:
        raise ValueError("Table 1 'weights' header row not found")
    hdr = rows[hdr_i]
    blocks = [i for i, c in enumerate(hdr)
              if isinstance(c, str) and c.strip().lower() == "weights"]
    if len(blocks) != 3:
        raise ValueError(f"expected 3 geography blocks, found {len(blocks)} at {blocks}")

    # report period from the current-index column header of the first block, 'M/YYYY'
    m = re.search(r"(\d{1,2})/(\d{4})", str(hdr[blocks[0] + 1]))
    if not m:
        raise ValueError(f"could not read period from header {hdr[blocks[0] + 1]!r}")
    period = f"{m.group(2)}-{int(m.group(1)):02d}"

    # confirm the block order really is Urban, Rural, Total (fail loud otherwise)
    geo_row = next((r for r in rows
                    if any(isinstance(c, str) and "urban egypt" in c.lower() for c in r)), None)
    order = [c.strip().split()[0].title()
             for c in (geo_row or []) if isinstance(c, str) and c.strip().lower().endswith("egypt")]
    if order != [g for g, _ in _GEOS]:
        raise ValueError(f"unexpected geography order in Table 1: {order}")
    label_col = len(hdr) - 1  # English label is the last column

    # pick the first row matching each of the 13 divisions
    picked: dict[str, tuple[str, list]] = {}
    for r in rows:
        lab = r[label_col] if label_col < len(r) else None
        if not isinstance(lab, str) or not lab.strip():
            continue
        s = _norm(lab)
        is_upper = lab == lab.upper()          # sub-groups are Title Case
        for code, clabel, rule in _DIV_RULES:
            if code in picked:
                continue
            if (is_upper or code == "00") and rule(s):
                picked[code] = (clabel, r)
                break
    missing = [c for c, _, _ in _DIV_RULES if c not in picked]
    if missing:
        raise ValueError(f"Table 1 missing division rows: {missing}")

    records = []
    for code, _, _ in _DIV_RULES:
        clabel, r = picked[code]
        for (geo, _tok), w in zip(_GEOS, blocks):
            idx, mom, yoy = r[w + 1], r[w + 2], r[w + 3]
            if idx is None:
                continue
            records.append((code, clabel, geo, period, "index", round(float(idx), 4), "Index", _BASE_PERIOD))
            if mom is not None:
                records.append((code, clabel, geo, period, "inflation_mom", round(float(mom), 4), "percent", ""))
            if yoy is not None:
                records.append((code, clabel, geo, period, "inflation_yoy", round(float(yoy), 4), "percent", ""))

    out = pd.DataFrame.from_records(
        records,
        columns=["coicop_code", "coicop_label", "geography", "period", "measure",
                 "value", "unit", "base_period"],
    )
    out["frequency"] = "monthly"
    return out
