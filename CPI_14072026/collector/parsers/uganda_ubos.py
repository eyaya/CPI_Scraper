"""Parser for the Uganda UBOS monthly 'CPI Excel Tables' workbook (Tier 2).

UBOS publishes a clean wide time series each month. The 'Division' sheet stacks
six blocks, each a wide table with one column per month (Jul 2017 -> latest):

  National, by COICOP-2018 division:
    * index levels        (rows: 13 divisions + 'Grand Total' = All items)
    * 'Annual% Change'    (YoY %,  'Headline' + 13 divisions)
    * 'Monthly % Change'  (MoM %,  13 divisions)
  By urban centre (Kampala income tiers, Masaka, Mbarara, ... Fortportal):
    * index / 'Annual % Change' / 'Monthly % Change'  (All-items per centre)

Each block is introduced by a date-header row (many datetime cells) whose label
gives the measure ('Annual'->yoy, 'Monthly'->mom, else index); the centre index
block is marked by a 'Centre' row instead. Uganda uses full COICOP-2018, so the
division number (1..13) is the code directly and 'Grand Total'/'Headline' is All
items (00). We emit index + inflation_yoy + inflation_mom for the national
divisions and for each centre (All-items), de-duplicating the national All-items
that appears in more than one block. Base 2016/17 = 100.
"""
from __future__ import annotations
import datetime as dt
import openpyxl
import pandas as pd

from ..coicop import DIVISIONS

_BASE_PERIOD = "2016/17 = 100"
# label cells that denote the national All-items headline (not a named centre)
_NATIONAL_00 = {"grand total", "headline", "centre"}


def _is_date_header(row) -> bool:
    return sum(isinstance(c, (dt.datetime, dt.date)) for c in row) >= 12


def _month_map(row) -> dict[int, str]:
    out = {}
    for i, c in enumerate(row):
        if isinstance(c, (dt.datetime, dt.date)):
            out[i] = f"{c.year}-{c.month:02d}"   # day component is a placeholder
    return out


def _measure_of(row) -> str:
    text = " ".join(str(c) for c in row if isinstance(c, str)).lower()
    if "annual" in text:
        return "inflation_yoy"
    if "monthly" in text:
        return "inflation_mom"
    return "index"


def parse(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == "division"), None)
    if sheet is None:
        raise ValueError(f"'Division' sheet not found in {wb.sheetnames}")
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]

    months: dict[int, str] = {}
    measure = "index"
    records = []

    for r in rows:
        if _is_date_header(r):
            months = _month_map(r)
            measure = _measure_of(r)
            # a date-header may also be its first data row (e.g. 'Headline')
        if not months:
            continue
        code = label = geography = None
        num = r[0]
        lab = r[1].strip() if isinstance(r[1], str) else None
        if isinstance(num, (int, float)) and 1 <= int(num) <= 13:
            code = f"{int(num):02d}"
            label = str(r[1]).strip() if r[1] is not None else DIVISIONS.get(code, "")
            geography = "National"
        elif lab and lab.lower() in _NATIONAL_00:
            code, label, geography = "00", "All items", "National"
            if lab.lower() == "centre":       # 'Centre' row starts the centre index block
                measure = "index"
        elif lab:                             # a named urban centre, All-items
            code, label, geography = "00", "All items", lab
        else:
            continue

        for col, period in months.items():
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)):
                records.append((code, label, geography, period, measure,
                                round(float(v), 4)))

    out = pd.DataFrame.from_records(
        records,
        columns=["coicop_code", "coicop_label", "geography", "period", "measure", "value"],
    )
    # national All-items appears in several blocks (Grand Total / Centre / Headline)
    out = out.drop_duplicates(["coicop_code", "geography", "period", "measure"])
    out["unit"] = out["measure"].map(lambda m: "Index" if m == "index" else "percent")
    out["base_period"] = out["measure"].map(lambda m: _BASE_PERIOD if m == "index" else "")
    out["frequency"] = "monthly"
    return out
