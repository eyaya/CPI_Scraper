"""Parser for the Namibia NSA monthly 'CPI Excel Tables' workbook (Tier 2).

NSA publishes a rich workbook each month. Three sheets are wide monthly time
series (from 2002, base Dec. 2012 = 100) by COICOP-1999 division:

  * 'Tab 2' — index levels
  * 'Tab 3' — month-on-month % change
  * 'Tab 4' — year-on-year % change

Each has: <code> | <label> | <one column per month>. Division rows carry a
2-digit code (00 = All items, 01..12); the interleaved sub-groups (Food, Bread
and cereals, …) have a blank code and are skipped. We emit index + inflation_mom
+ inflation_yoy by division for the full history. Labels kept as published.
"""
from __future__ import annotations
import datetime as dt
import re
import openpyxl
import pandas as pd

_BASE_PERIOD = "Dec. 2012 = 100"
# (sheet-name prefix, measure, unit, base_period)
_SHEETS = [
    ("Tab 2", "index", "Index", _BASE_PERIOD),
    ("Tab 3", "inflation_mom", "percent", ""),
    ("Tab 4", "inflation_yoy", "percent", ""),
]


def _code(raw) -> str | None:
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)) and float(raw).is_integer() and 0 <= raw <= 12:
        return f"{int(raw):02d}"
    if isinstance(raw, str):
        s = raw.strip().rstrip(".").strip()
        if re.fullmatch(r"\d{1,2}", s) and 0 <= int(s) <= 12:
            return f"{int(s):02d}"
    return None


def _extract(ws, measure, unit, base) -> list[tuple]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows)
               if sum(isinstance(c, (dt.datetime, dt.date)) for c in r) >= 12), None)
    if hi is None:
        return []
    months = {j: f"{c.year}-{c.month:02d}" for j, c in enumerate(rows[hi])
              if isinstance(c, (dt.datetime, dt.date))}
    out = []
    for row in rows[hi + 1:]:
        code = _code(row[0] if row else None)
        if code is None:                       # sub-groups have a blank code
            continue
        label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        for j, period in months.items():
            v = row[j] if j < len(row) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append((code, label, period, measure, round(float(v), 4), unit, base))
    return out


def parse(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []
    for prefix, measure, unit, base in _SHEETS:
        sheet = next((s for s in wb.sheetnames if s.strip() == prefix), None)
        if sheet is not None:
            records += _extract(wb[sheet], measure, unit, base)
    if not records:
        raise ValueError("no CPI division series found in NSA workbook")

    out = pd.DataFrame.from_records(
        records,
        columns=["coicop_code", "coicop_label", "period", "measure", "value",
                 "unit", "base_period"])
    out = out.drop_duplicates(["coicop_code", "period", "measure"])
    out["geography"] = "National"
    out["frequency"] = "monthly"
    return out
